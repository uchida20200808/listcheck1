#!/usr/bin/env python
# coding: utf-8

# In[66]:


import pandas as pd
import numpy as np
import glob
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles.alignment import Alignment
import datetime
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta


# In[67]:


files = glob.glob('*更新リスト*.xlsx')
for file in files:
    print(file + 'を読み込みました。')


# In[68]:


okyakusama = pd.read_excel(file) #契約一覧の読み込み
#okyakusama_pick = okyakusama[['契約者名', '契約者名カナ','保険会社(名称)','保険種目(名称)','証券番号','満期日',郵便番号','契約者住所']]
#okyakusama['お客様氏名（漢字）'] = okyakusama['お客様氏名（漢字）'].str.replace('　', '')
okyakusama.head()


# In[69]:


okyakusama['超保険・長火'] = 'False'


# In[70]:


'''
同姓同名の処理が課題
'''


# In[71]:


okyakusama.head()


# In[72]:


files2 = glob.glob('*wise*.xlsx')
for file2 in files2:
    print(file2 + 'を読み込みました。')


# In[73]:


wb = load_workbook(file2)
ws = wb.active


# In[74]:


row_list = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                       min_col=1, max_col=ws.max_column):
    cell_list = [cell.value for cell in row]
    row_list.append(cell_list) #エクセルのデータをプログラムで読み込める用に処理する


# In[ ]:





# In[75]:


for i in row_list:
    list_name = str(i[2]).replace('　', '') #契約者名
    #okyakusama['お客様氏名（漢字）'] = okyakusama['お客様氏名（漢字）'].str.replace('　', '')
    list_syokenNO = i[40] #証券番号
    
    okyakusama.loc[(okyakusama['契約者'] == list_name) & (okyakusama['旧証券番号'] == list_syokenNO), '超保険・長火'] = 'True'


# In[76]:


print(okyakusama.loc[okyakusama['超保険・長火'] == 'False'])


# In[84]:


checked_data = okyakusama.loc[okyakusama['超保険・長火'] == 'False']
manki = pd.to_timedelta(checked_data['満期日'],unit='D')+pd.to_datetime("1899/12/30")
checked_data['満期日'] = manki
checked_data = checked_data.sort_values('満期日', ascending=True)


# In[ ]:





# In[ ]:





# In[ ]:





# In[85]:


checked_data.head()


# In[86]:


checked_data.to_excel('checked_data.xlsx',index=False)


# In[ ]:





# In[ ]:




